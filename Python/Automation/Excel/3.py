import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, Border 
from openpyxl.styles import Side, PatternFill, Color
from openpyxl.styles import GradientFill, colors, NamedStyle
loc = 'C:/Users/allen/OneDrive/Desktop/Github/Learning/' \
      'Python/Automation/Excel/data/2'                          # Location of the directory
wb = Workbook()                                                 # Create workbook
ws = wb.active                                                  # Go to active sheet of 

for z in range(1,20): ws.append(range(300))

ws.merge_cells('B2:E5')                                         # Merge cells B2:E5
ws.unmerge_cells('B2:E5')                                       # Unmerge cells B2:E5
ws.merge_cells(
    start_row=2, start_column=2, 
    end_row=5, end_column=5
)                                                               # Merge cells B2:E5 

# Merged Cell Formatting
cell = ws['B2']                                                 # Get cell B2
cell.value = 'Allen'                                            # Set value
cell.font = Font(
    name='Calibri', color=colors.COLOR_INDEX[44],
    size=24, bold=True
)                                                               # Set font
cell.alignment = Alignment(
    horizontal='center', vertical='center'
)                                                               # Set alignment
cell.fill = GradientFill(
    stop=('009FFD', '2A2A72'), type='linear'
)                                                               # Set gradient fill

# Highlight Style 1
h1 = NamedStyle(name='highlight1')                              # Create a style
h1.font = Font(bold=True)                                       # Set font
bd1 = Side(style='thick', color='FFFF00')                       # Set border
h1.fill = PatternFill('solid', fgColor='FFFF00')                # Set fill
h1.border = Border(
    left=bd1, top=bd1, 
    right=bd1, bottom=bd1
)                                                               # Apply border

# Highlight Style 2
h2 = NamedStyle(name='highlight2')                              # Create a style
h2.font = Font(italic=True)                                     # Set font
bd2 = Side(style='thin', color='000000')                        # Set border
h2.fill = GradientFill(
    stop=('83EAF1', '63A4FF'), type='linear'
)                                                               # Set fill
h2.border = Border(
    left=bd2, top=bd2, 
    right=bd2, bottom=bd2
)                                                               # Apply border

# Apply highlights
z = 0 
for col in ws.iter_cols(
    min_row=1, min_col=7, 
    max_col=9, max_row=10
):
    col[z+1].style = h1
    col[z+2].style = h2
    z+=1

wb.save(f'{loc}/B_Write.xlsx')                                  # Save workbook