import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
loc = 'C:/Users/allen/OneDrive/Desktop/Github/Learning/' \
      'Python/Automation/Excel/data/3'                            # Location of the directory

# Merge Dataframes
df1 = pd.read_excel(f'{loc}/A_Read.xlsx', sheet_name='Sheet1')    # Read the first file sheet 1
df2 = pd.read_excel(f'{loc}/A_Read.xlsx', sheet_name='Sheet2')    # Read the first file sheet 2        
df3 = pd.read_excel(f'{loc}/B_Read.xlsx')                         # Read the second file
df = pd.concat([df1, df2, df3], sort=False)                       # Combine the dataframes
df.to_excel(f'{loc}/D_Write.xlsx')                                # Write the dataframe to a file

# Modify Workbooks
wb = load_workbook(f'{loc}/D_Write.xlsx')                         # Load the file
ws = wb.active                                                    # Get the active sheet
col = ws['H1']                                                    # Get the column
col.font = Font(bold=True)                                        # Set the font to bold
col.value = 'Total'                                               # Set the value

for z in range(2,300):                                            # Loop through the rows
      f = ws['F'+str(z)].value                                    # Value of cell in E column
      g = ws['G'+str(z)].value                                    # Value of cell in F column
      ws['H'+str(z)].value = g + f                                # Save the calculated value

rows = dataframe_to_rows(df, index=False, header=True)            # Convert the dataframe to rows
for z in rows: print(z)                                           # Print the rows
for z,r1 in enumerate(rows,1):                                      # Loop through the rows
      for y,r2 in rows:                                             # Loop through the columns
            if r1[0] == r2[0]:                                    # If the first column is the same

wb.save(f'{loc}/D_Write.xlsx')                                    # Save the file